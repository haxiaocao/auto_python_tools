# -*- coding: utf-8 -*-
# reference:https://automatetheboringstuff.com/chapter12/
import openpyxl

file_name = 'example.xlsx'
sheet_name = 'rename1'


def modify_fruit_price(file_name, sheet_name):
    fruit_price_dic = {'Apples': 100, 'Pears': 200, 'Bananas': 300}
    wb = openpyxl.load_workbook(file_name)
    sheet = wb.get_sheet_by_name(sheet_name)
    # print sheet.title

    for i in range(2, sheet.max_row):
        fruit_cell = sheet.cell(row=i, column=2)
        fruit_name = fruit_cell.value
        if fruit_name in fruit_price_dic:
            price_cell = sheet.cell(row=i, column=3)
            price_cell.value = fruit_price_dic[fruit_name]

    wb.save(file_name)


def modidy_cell_styles():
    from openpyxl.styles import Font
    from openpyxl.styles import Alignment
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet['A1'] = 'Hello, this world.'
    sheet['A2'] = 'New World'
    sheet['A1'].font = Font(bold=True, italic=True, color='FF0000')
    # set the row height and column width: their measurement unit differs.
    sheet.row_dimensions[1].height = 50
    sheet.column_dimensions['A'].width = 30

    sheet.merge_cells('C5:F7')
    sheet['C5'] = 'How do you do ?'
    sheet.merge_cells('I2:I5')
    sheet['I2'] = 'Merge Rows from here and there .'

    # alignment the text in the cells or merged cells.
    sheet['C5'].alignment = Alignment(horizontal='center', vertical='center')

    wb.save('examaple2.xlsx')


# modify_fruit_price(file_name, sheet_name)
modidy_cell_styles()
